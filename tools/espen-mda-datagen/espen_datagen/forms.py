"""Compile ESPEN XLSForms to XForms (pyxform) and extract fillable schemas."""
from __future__ import annotations

import copy
import os
from dataclasses import dataclass, field

from lxml import etree
import openpyxl
from pyxform.xls2xform import convert

XF = "http://www.w3.org/2002/xforms"
NS = {"h": "http://www.w3.org/1999/xhtml", "x": XF}

FORM_KEYS = [
    "1_location",
    "2_part",
    "3_med_treatment",
    "4_case_mngnt",
    "5_supervision_hf",
    "6_supervision_CDD",
]


@dataclass
class FormSchema:
    key: str
    form_id: str
    version: str | None
    title: str
    template: etree._Element  # deep copy of the <data> instance node
    leaf_names: list[str]
    calculate_names: set[str]
    choices: dict[str, list[str]]
    src_path: str = ""
    meta_present: bool = False
    xform_xml: str = ""  # compiled XForm, persisted for Ona Data publishing


def form_path(forms_dir: str, key: str) -> str:
    return os.path.join(forms_dir, f"demo_mda_9999_{key}.xlsx")


def _localname(el) -> str:
    tag = el.tag
    return tag.split("}")[-1] if isinstance(tag, str) else tag


def _collect_leaves(data_el) -> tuple[list[str], bool]:
    """Localnames of value-bearing leaves, in doc order, excluding the meta subtree."""
    names: list[str] = []
    meta_present = False
    for child in data_el.iter():
        if child is data_el:
            continue
        if _localname(child) == "meta":
            meta_present = True
    for child in data_el.iter():
        if child is data_el:
            continue
        ln = _localname(child)
        # skip anything under meta
        parent = child.getparent()
        in_meta = False
        while parent is not None and parent is not data_el:
            if _localname(parent) == "meta":
                in_meta = True
                break
            parent = parent.getparent()
        if in_meta or ln == "meta":
            continue
        if len(child) == 0:  # leaf
            names.append(ln)
    return names, meta_present


def _calculate_names(root) -> set[str]:
    out = set()
    for b in root.findall(".//x:bind", NS):
        if b.get("calculate"):
            ref = b.get("nodeset", "")
            out.add(ref.rsplit("/", 1)[-1])
    return out


def _load_choices(xlsx_path: str) -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "choices" not in wb.sheetnames:
        return {}
    ws = wb["choices"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(h) if h is not None else "" for h in rows[0]]
    li = header.index("list_name") if "list_name" in header else 0
    ni = header.index("name") if "name" in header else 1
    out: dict[str, list[str]] = {}
    for r in rows[1:]:
        if not r or li >= len(r) or r[li] is None:
            continue
        name = r[ni] if ni < len(r) and r[ni] is not None else None
        if name is None:
            continue
        out.setdefault(str(r[li]).strip(), []).append(str(name).strip())
    return out


def load_form(xlsx_path: str) -> FormSchema:
    result = convert(xlsx_path)
    root = etree.fromstring(result.xform.encode("utf-8"))
    instance = root.find(".//x:model/x:instance", NS)
    data_el = list(instance)[0]
    form_id = data_el.get("id")
    version = data_el.get("version")
    title_el = root.find(".//h:head/h:title", NS)
    title = title_el.text if title_el is not None else form_id
    leaves, meta_present = _collect_leaves(data_el)
    key = "_".join(os.path.basename(xlsx_path).replace(".xlsx", "").split("_")[2:])
    return FormSchema(
        key=key,
        form_id=form_id,
        version=version,
        title=title,
        template=copy.deepcopy(data_el),
        leaf_names=leaves,
        calculate_names=_calculate_names(root),
        choices=_load_choices(xlsx_path),
        src_path=xlsx_path,
        meta_present=meta_present,
        xform_xml=result.xform,
    )


def load_all(forms_dir: str) -> dict[str, FormSchema]:
    return {k: load_form(form_path(forms_dir, k)) for k in FORM_KEYS}
